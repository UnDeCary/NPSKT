from io import BytesIO
import re
from typing import Any

from openpyxl import load_workbook


CANONICAL_COLUMNS = {
    "ID анкеты",
    "Дата интервью",
    "Период / волна",
    "Сегмент",
    "База",
    "Компания",
    "Продукт",
    "Технология",
    "Макрорегион",
    "Регион / область",
    "Населенный пункт",
    "Тип населенного пункта",
    "Оценка NPS",
}

COLUMN_ALIASES = {
    "ID Анкеты": "ID анкеты",
    "ID анкеты": "ID анкеты",
    "Дата интервью": "Дата интервью",
    "Период / волна": "Период / волна",
    "ВОЛНА": "Период / волна",
    "Волна": "Период / волна",
    "Сегмент": "Сегмент",
    "База": "База",
    "Компания": "Компания",
    "Услуга": "Продукт",
    "Продукт": "Продукт",
    "Технология": "Технология",
    "Макрорегион": "Макрорегион",
    "Регион / область": "Регион / область",
    "Населенный пункт": "Населенный пункт",
    "Тип населенного пункта": "Тип населенного пункта",
    "Оценка": "Оценка NPS",
    "Оценка NPS": "Оценка NPS",
}

COMPANY_MAP = {
    "Kazahtelecom": "Казахтелеком",
    "Kaztelecom": "Казахтелеком",
    "Казахтелеком": "Казахтелеком",
    "Transtelekom": "Транстелеком",
    "Транстелеком": "Транстелеком",
    "AlmaTV": "Alma TV",
    "Alma TV": "Alma TV",
    "Beeline": "Beeline",
    "Jusan": "Jusan",
    "Kazteleport": "Казахтелеком",
}

PRODUCT_MAP = {
    "Интернет": "Интернет",
    "TV": "Телевидение",
    "Телевидение": "Телевидение",
    "FMS": "FMS",
    "ИКТ": "ИКТ / Хостинг",
    "ИКТ / Хостинг": "ИКТ / Хостинг",
    "Облачное видео": "Облачное видеонаблюдение",
    "Облачное видеонаблюдение": "Облачное видеонаблюдение",
}

REGION_META = {
    "Астана": ("Астана", "Астана", "Астана"),
    "Алматы": ("Алматы", "Алматы", "Алматы"),
    "Акмолинская область (Кокшетау)": ("Север", "Акмолинская область", "Кокшетау"),
    "Алматинская область (Конаев)": ("Алматы", "Алматинская область", "Конаев"),
    "Абайская область (Семей)": ("Восток", "Абайская область", "Семей"),
    "Актюбинская область (Актобе)": ("Запад", "Актюбинская область", "Актобе"),
    "Атырауская область (Атырау)": ("Запад", "Атырауская область", "Атырау"),
    "Восточно-Казахстанская область (Усть-Каменогорск)": ("Восток", "Восточно-Казахстанская область", "Усть-Каменогорск"),
    "Западно-Казахстанская область (Уральск)": ("Запад", "Западно-Казахстанская область", "Уральск"),
    "Жамбылская область (Тараз)": ("Юг", "Жамбылская область", "Тараз"),
    "Жетысуская область (Талдыкурган)": ("Алматы", "Жетысуская область", "Талдыкорган"),
    "Карагандинская область (Караганда)": ("Центр", "Карагандинская область", "Караганда"),
    "Костанайская область (Костанай)": ("Север", "Костанайская область", "Костанай"),
    "Кызылординская область (Кызылорда)": ("Юг", "Кызылординская область", "Кызылорда"),
    "Мангистауская область (Актау)": ("Запад", "Мангистауская область", "Актау"),
    "Павлодарская область (Павлодар)": ("Север", "Павлодарская область", "Павлодар"),
    "Северо-Казахстанская область (Петропавловск)": ("Север", "Северо-Казахстанская область", "Петропавловск"),
    "Улытауская область (Жезказган)": ("Центр", "Улытауская область", "Жезказган"),
    "Шымкент и Туркестанская область (Туркестан)": ("Юг", "Туркестанская область", "Шымкент"),
}


def _clean(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _score(value: Any) -> int | Any:
    if isinstance(value, int):
        return value
    match = re.search(r"\d+", _clean(value))
    return int(match.group(0)) if match else value


def _wave(value: Any) -> str:
    text = _clean(value)
    return {
        "H1 2026": "I полугодие 2026",
        "H2 2026": "II полугодие 2026",
    }.get(text, text)


def _normalize_company(value: Any) -> str:
    text = _clean(value)
    return COMPANY_MAP.get(text, text)


def _normalize_product(value: Any) -> str:
    text = _clean(value)
    return PRODUCT_MAP.get(text, text)


def _normalize_sample_layout(row: dict[str, Any], headers: list[str]) -> dict[str, Any]:
    company = _normalize_company(row.get("Компания"))
    product = _normalize_product(row.get("Услуга"))
    technology = _clean(row.get("Продукт"))
    if technology == "-":
        technology = ""

    region_header = _clean(row.get("Регион B2B"))
    settlement_type = ""
    if not region_header or region_header == "-":
        for header in headers:
            value = _clean(row.get(header))
            if header in REGION_META and value:
                region_header = header
                settlement_type = value if value in {"Город", "Село"} else ""
                break

    macroregion, region, settlement = REGION_META.get(region_header, ("Центр", region_header, region_header))
    segment = _clean(row.get("Сегмент"))
    if segment == "B2B":
        product = {
            "Интернет": "Корпоративный интернет",
            "ИКТ": "ИКТ / Хостинг",
            "Облачное видео": "Облачное видеонаблюдение",
        }.get(_clean(row.get("Услуга")), product)

    return {
        "ID анкеты": _clean(row.get("ID Анкеты") or row.get("ID анкеты")),
        "Дата интервью": _clean(row.get("Дата интервью")) or "2026-01-01",
        "Период / волна": _wave(row.get("ВОЛНА") or row.get("Волна") or row.get("Период / волна")),
        "Сегмент": segment,
        "База": "Казахтелеком" if company == "Казахтелеком" else "конкуренты",
        "Компания": company,
        "Продукт": product,
        "Технология": technology,
        "Макрорегион": macroregion,
        "Регион / область": region,
        "Населенный пункт": settlement,
        "Тип населенного пункта": settlement_type,
        "Оценка NPS": _score(row.get("Оценка") or row.get("Оценка NPS")),
    }


def _normalize_canonical(row: dict[str, Any]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for key, value in row.items():
        normalized[COLUMN_ALIASES.get(key, key)] = value
    if "Период / волна" in normalized:
        normalized["Период / волна"] = _wave(normalized["Период / волна"])
    if "Компания" in normalized:
        normalized["Компания"] = _normalize_company(normalized["Компания"])
    if "Продукт" in normalized:
        normalized["Продукт"] = _normalize_product(normalized["Продукт"])
    if "Оценка NPS" in normalized:
        normalized["Оценка NPS"] = _score(normalized["Оценка NPS"])
    return normalized


def parse_xlsx(content: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_clean(value) for value in rows[0]]
    parsed = []
    is_sample_layout = {"ВОЛНА", "Услуга", "ID Анкеты", "Оценка", "NPS"}.issubset(set(headers))
    for row in rows[1:]:
        item = {headers[index]: value for index, value in enumerate(row) if index < len(headers) and headers[index]}
        if not any(value is not None and str(value).strip() != "" for value in item.values()):
            continue
        parsed.append(_normalize_sample_layout(item, headers) if is_sample_layout else _normalize_canonical(item))
    return parsed
